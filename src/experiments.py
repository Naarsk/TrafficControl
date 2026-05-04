"""Experiment orchestration for the F4C2 traffic-MDP homework.

Replicates Table 1 (symmetric F4C2) of Haijema & van der Wal (2008)
using Value Iteration and Policy Iteration.

Results from every run are persisted to ``results/<timestamp>.json``
(and an accompanying ``.xlsx``) so they can be revisited, plotted, or
compared across configurations without having to re-solve.
"""

from __future__ import annotations

import json
import platform
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path

import numpy as np

from .model import F4C2Model, F4C2Params
from .solvers import value_iteration, policy_iteration


# Paper Table 1, MDP-optimal cyclic E[W] in seconds (symmetric F4C2):
PAPER_TABLE1: dict[float, float] = {
    0.40: 4.89,
    0.60: 6.95,
    0.80: 13.5,
}


def little_law_E_W(g: float, q: tuple[float, ...], slot_seconds: float = 2.0) -> float:
    """Convert avg cars per slot (g) to E[W] in seconds.

    By Little's law L = lambda * W, with L = g (avg cars in system) and
    lambda = total arrival rate. Arrival rate is sum_f q_f cars per slot,
    or sum_f q_f / slot_seconds cars per second.
    """
    arrivals_per_second = sum(q) / slot_seconds
    return g / arrivals_per_second


def q_for_rho(rho: float) -> tuple[float, float, float, float]:
    """Per-flow Bernoulli arrival prob giving symmetric workload rho.

    For F4C2 with C_1 = {1, 3}, C_2 = {2, 4}, the workload is
        rho = max_{f in C_1} q_f + max_{f in C_2} q_f.
    Symmetric case (all q_f equal): q_f = rho / 2.
    """
    q_each = rho / 2.0
    return (q_each, q_each, q_each, q_each)


@dataclass
class WorkloadResult:
    rho: float
    K: int
    g_vi: float | None = None
    g_pi: float | None = None
    e_w_vi: float | None = None
    e_w_pi: float | None = None
    iters_vi: int | None = None
    iters_pi: int | None = None
    time_build: float = 0.0
    time_vi: float = 0.0
    time_pi: float = 0.0
    pi_failed: bool = False
    pi_error: str | None = None


RESULTS_DIR = Path("results")


def save_results_json(
    results: list[WorkloadResult],
    args_dict: dict,
    results_dir: Path = RESULTS_DIR,
    tag: str = "run",
) -> Path:
    """Persist a list of WorkloadResult to results/<timestamp>_<tag>.json."""
    results_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = results_dir / f"{ts}_{tag}.json"
    payload = {
        "timestamp": ts,
        "tag": tag,
        "args": args_dict,
        "platform": {
            "python": platform.python_version(),
            "system": platform.system(),
            "machine": platform.machine(),
        },
        "results": [asdict(r) for r in results],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    print(f"\n  [persist] results saved to {path}")
    return path


def save_results_xlsx(
    results: list[WorkloadResult],
    args_dict: dict,
    results_dir: Path = RESULTS_DIR,
    tag: str = "run",
) -> Path:
    """Persist results to results/<timestamp>_<tag>.xlsx (3 sheets).

    Sheets:
      * "summary" - one row per workload, key metrics + paper deltas + cross-checks
      * "raw"     - full WorkloadResult dataclass dump
      * "args"    - the CLI arguments and platform info
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    results_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = results_dir / f"{ts}_{tag}.xlsx"

    wb = Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")

    # ---- summary sheet ----
    ws = wb.active
    ws.title = "summary"
    cols = [
        ("rho", "rho"),
        ("paper E[W] (s)", "paper_e_w"),
        ("E[W]_VI (s)", "e_w_vi"),
        ("E[W]_PI (s)", "e_w_pi"),
        ("g_VI", "g_vi"),
        ("g_PI", "g_pi"),
        ("|g_VI - g_PI|", "_gap_vi_pi"),
        ("delta_VI vs paper (s)", "_delta_vi"),
        ("PI status", "_pi_status"),
        ("K", "K"),
        ("iters_VI", "iters_vi"),
        ("iters_PI", "iters_pi"),
        ("time_VI (s)", "time_vi"),
        ("time_PI (s)", "time_pi"),
    ]
    for j, (header, _) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(results, start=2):
        derived = {
            "paper_e_w": PAPER_TABLE1.get(r.rho),
            "_gap_vi_pi": (
                abs(r.g_vi - r.g_pi)
                if (r.g_vi is not None and r.g_pi is not None)
                else None
            ),
            "_delta_vi": (
                r.e_w_vi - PAPER_TABLE1[r.rho]
                if (r.rho in PAPER_TABLE1 and r.e_w_vi is not None)
                else None
            ),
            "_pi_status": "FAIL" if r.pi_failed else ("ok" if r.g_pi is not None else "skip"),
        }
        row = asdict(r)
        row.update(derived)
        for j, (_, key) in enumerate(cols, start=1):
            value = row.get(key)
            ws.cell(row=i, column=j, value=value)

    for j, (header, _) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(len(header) + 2, 12)
    ws.freeze_panes = "A2"

    # ---- raw sheet ----
    raw = wb.create_sheet("raw")
    if results:
        keys = list(asdict(results[0]).keys())
        for j, k in enumerate(keys, start=1):
            cell = raw.cell(row=1, column=j, value=k)
            cell.font = header_font
            cell.fill = header_fill
        for i, r in enumerate(results, start=2):
            d = asdict(r)
            for j, k in enumerate(keys, start=1):
                raw.cell(row=i, column=j, value=d[k])
        for j, k in enumerate(keys, start=1):
            raw.column_dimensions[get_column_letter(j)].width = max(len(k) + 2, 12)
        raw.freeze_panes = "A2"

    # ---- args sheet ----
    args_ws = wb.create_sheet("args")
    args_ws.cell(row=1, column=1, value="key").font = header_font
    args_ws.cell(row=1, column=2, value="value").font = header_font
    args_ws.cell(row=1, column=1).fill = header_fill
    args_ws.cell(row=1, column=2).fill = header_fill
    rows = [
        ("timestamp", ts),
        ("tag", tag),
        ("python", platform.python_version()),
        ("system", platform.system()),
        ("machine", platform.machine()),
    ]
    for k, v in args_dict.items():
        rows.append((k, str(v) if v is not None else ""))
    for i, (k, v) in enumerate(rows, start=2):
        args_ws.cell(row=i, column=1, value=k)
        args_ws.cell(row=i, column=2, value=v)
    args_ws.column_dimensions["A"].width = 16
    args_ws.column_dimensions["B"].width = 40

    wb.save(path)
    print(f"  [persist] xlsx saved to {path}")
    return path


def _algo_set(algo: str) -> set[str]:
    """Resolve a CLI --algo value into a set of solvers to run."""
    if algo in {"both", "vi-pi", "vi+pi", "all"}:
        return {"vi", "pi"}
    if algo in {"vi", "pi"}:
        return {algo}
    raise ValueError(f"unknown --algo {algo!r}")


def run_single(
    rho: float,
    K: int = 10,
    algo: str = "both",
    tol: float = 1e-6,
    log_interval: float = 1.0,
    persist: bool = True,
) -> WorkloadResult:
    """Build the model at workload rho and solve with the chosen algorithm(s)."""
    q = q_for_rho(rho)
    params = F4C2Params(q=q, K=K)
    res = WorkloadResult(rho=rho, K=K)
    which = _algo_set(algo)

    print(f"\n=== rho = {rho:.2f}    q_f = {q[0]:.3f}    K = {K}    tau = {params.tau} ===")

    t0 = time.time()
    model = F4C2Model(params)
    res.time_build = time.time() - t0

    if "vi" in which:
        print(f"\n--- Value Iteration (K={K}) ---")
        vi = value_iteration(model, tol=tol, log_interval=log_interval)
        res.g_vi = vi.g
        res.iters_vi = vi.iterations
        res.time_vi = vi.elapsed
        res.e_w_vi = little_law_E_W(vi.g, q, params.slot_seconds)
        ps = model.policy_summary(vi.policy_sa)
        print(
            f"  -> g_VI = {vi.g:.6f}    E[W]_VI = {res.e_w_vi:.3f} s    "
            f"iters = {vi.iterations}    "
            f"second-action picks = {ps['n_pick_second_action']}/{ps['n_decision_states']}"
        )

    if "pi" in which:
        print(f"\n--- Policy Iteration (K={K}) ---")
        try:
            pi = policy_iteration(model, log_interval=log_interval)
            res.g_pi = pi.g
            res.iters_pi = pi.iterations
            res.time_pi = pi.elapsed
            res.e_w_pi = little_law_E_W(pi.g, q, params.slot_seconds)
            ps = model.policy_summary(pi.policy_sa)
            print(
                f"  -> g_PI = {pi.g:.6f}    E[W]_PI = {res.e_w_pi:.3f} s    "
                f"iters = {pi.iterations}    "
                f"second-action picks = {ps['n_pick_second_action']}/{ps['n_decision_states']}"
            )
        except RuntimeError as exc:
            res.pi_failed = True
            res.pi_error = str(exc)
            print(f"  -> PI did not converge: {exc}")

    # Cross-check
    print()
    if res.g_vi is not None and res.g_pi is not None:
        print(f"  cross-check |g_VI - g_PI|@K={K}        = {abs(res.g_vi - res.g_pi):.3e}")

    if rho in PAPER_TABLE1:
        target = PAPER_TABLE1[rho]
        print(f"  paper Table 1 target E[W] = {target} s")
        if res.e_w_vi is not None:
            print(f"    delta_VI@K={K}       = {res.e_w_vi - target:+.3f} s")
        if res.e_w_pi is not None:
            print(f"    delta_PI@K={K}       = {res.e_w_pi - target:+.3f} s")

    if persist:
        args_dict = {
            "cmd": "single",
            "rho": rho,
            "K": K,
            "algo": algo,
            "tol": tol,
        }
        tag = f"single_rho{rho:.2f}_K{K}"
        save_results_json([res], args_dict=args_dict, tag=tag)
        try:
            save_results_xlsx([res], args_dict=args_dict, tag=tag)
        except ImportError as exc:
            print(f"  [persist] xlsx skipped (openpyxl not installed): {exc}")

    return res


def _fmt(value: float | None, fmt: str = "9.3f") -> str:
    """Render a possibly-None float for the summary table."""
    if value is None:
        return f"{'--':>{int(fmt.split('.')[0])}}"
    return f"{value:{fmt}}"


def run_table1(
    K: int = 10,
    algo: str = "both",
    tol: float = 1e-6,
    log_interval: float = 1.0,
    persist: bool = True,
) -> list[WorkloadResult]:
    """Replicate Table 1: rho in {0.40, 0.60, 0.80} with identical arrival probs."""
    results = []
    for rho in (0.40, 0.60, 0.80):
        results.append(
            run_single(
                rho, K=K, algo=algo, tol=tol,
                log_interval=log_interval,
                persist=False,        # write a single combined file at the end instead
            )
        )

    width = 78
    print("\n" + "=" * width)
    print(f"Summary  (paper Table 1 MDP-optimal cyclic, symmetric F4C2; K={K})")
    print("=" * width)

    header = (
        f"{'rho':>5}  {'paper E[W]':>10}  "
        f"{'E[W]_VI':>9}  {'E[W]_PI':>9}  {'|g_VI-g_PI|':>13}"
    )
    print(header)
    print("-" * len(header))
    for r in results:
        target = PAPER_TABLE1.get(r.rho, float("nan"))
        ew_pi_str = "FAIL" if r.pi_failed else _fmt(r.e_w_pi)
        gap_vi_pi = (
            abs(r.g_vi - r.g_pi)
            if (r.g_vi is not None and r.g_pi is not None)
            else None
        )
        print(
            f"{r.rho:>5.2f}  {target:>10.3f}  "
            f"{_fmt(r.e_w_vi)}  {ew_pi_str:>9}  "
            f"{_fmt(gap_vi_pi, '13.2e')}"
        )
    print("=" * width)

    if persist:
        args_dict = {
            "cmd": "table1",
            "K": K,
            "algo": algo,
            "tol": tol,
        }
        save_results_json(results, args_dict, tag=f"table1_K{K}")
        try:
            save_results_xlsx(results, args_dict, tag=f"table1_K{K}")
        except ImportError as exc:
            print(f"  [persist] xlsx skipped (openpyxl not installed): {exc}")

    return results
